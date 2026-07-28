# DOH P2(테이크어웨이) 구현표 — KB P2 Feature 18개 × 되는것/안되는것/뷰/구현 VF
# 근거: kb/source/DOH_P2.xlsx Feature_Dictionary Category + 엔진 실제 게이팅.
# "P2는 클럽 없으면 통째로 불가"가 틀렸음을 기록 — 9/18은 몸 지표라 근사 P2로 측정(2026-07-28 구현).
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
FONT="맑은 고딕"

OK="✅"; NO="✕"; LO="◐"
# [Feature, 한글, 구분, 정면, 측면, 구현VF, 방법/왜 되나·안 되나]
R=[
("Head Sway Excessive (P2)","머리 과이동","🟢 몸(구현)",OK,NO,"VF152","머리 좌우이동 P1→P2 — VF031과 같은 식, 구간만 P2. 좌우축=정면 화면 안"),
("Pelvic Sway Excessive (P2)","골반 과이동","🟢 몸(구현)",OK,NO,"VF153","골반 좌우이동 P1→P2 — VF034 재사용"),
("Spine Angle Loss (P2)","척추각 손실","🟢 몸(구현)",NO,OK,"VF154","척추각 Δ P1→P2 음수 — VF038 재사용, 시상면=측면"),
("Spine Angle Increase (P2)","척추각 증가","🟢 몸(구현)",NO,OK,"VF154","같은 측정의 + 부호"),
("Trail Elbow Flexion Excessive (P2)","트레일 팔꿈치 조기굴곡","🟢 몸(구현)",OK,OK,"VF155","팔꿈치 내부각 @P2 — 세그먼트각이라 양쪽 뷰"),
("Lead Knee Flexion Excessive (P2)","리드 무릎 과굴곡","🟢 몸(구현)",OK,OK,"VF156","무릎 굽힘 Δ P1→P2 — VF039 재사용"),
("Hand Depth Deep (P2)","손 깊음","🟢 몸(구현)",NO,OK,"VF157","heel line 기준 — VF028과 같은 식, 시점만 P2. 앞뒤축=측면 화면 안"),
("Lead Knee Collapse (P2)","리드 무릎 안쪽 무너짐","🟢 몸(구현·신규 기하)",OK,NO,"VF158","힙-발목 선 대비 무릎의 스탠스축 이탈(다리길이 정규화) — 전두면=정면"),
("Body Rotation Dominant (P2)","몸통 회전 우세","🟢 몸(구현)",OK,OK,"VF013/016","회전 시계열 @P2 방출(Spec §2에 원래 있던 ID). '우세' 판정은 손이동과 조합 필요 — 재료 확보 단계"),
("Outside Takeaway (P2)","아웃사이드 테이크어웨이","🔴 클럽 필요",NO,NO,"—","클럽헤드 방향 — 포즈에 클럽이 없음. 2차 클럽엔진 몫"),
("Inside Takeaway (P2)","인사이드 테이크어웨이","🔴 클럽 필요",NO,NO,"—","상동"),
("Clubface Closed Excessive (P2)","클럽페이스 과닫힘","🔴 클럽 필요",NO,NO,"—","페이스각 — 클럽 검출+페이스 방향 필요"),
("Clubface Open Excessive (P2)","클럽페이스 과열림","🔴 클럽 필요",NO,NO,"—","상동"),
("Early Wrist Cock (P2)","얼리 코킹","🟡 손목(보류)",LO,LO,"—","마커리스 손 오차 8~20°(§5t 벤치마크) — 방출하면 오진 위험. 클럽 검출 후 샤프트각으로 대체 예정"),
("Late Wrist Cock (P2)","레이트 코킹","🟡 손목(보류)",LO,LO,"—","상동"),
("Lead Wrist Flexion Excessive (P2)","리드 손목 과굴곡","🟡 손목(보류)",LO,LO,"—","상동"),
("Lead Wrist Extension Excessive (P2)","리드 손목 과신전","🟡 손목(보류)",LO,LO,"—","상동"),
("Lead Pressure Excessive (P2)","리드측 압력 과다","🔴 압력판 필요",NO,NO,"—","압력(COP)은 영상으로 원리적 불가 — 무게중심(M801) 근사가 간접 대용 후보"),
]

wb=Workbook(); ws=wb.active; ws.title="P2 구현표 (18 Feature)"
bd=Border(*[Side(style="thin",color="DDE3E6")]*4)
head=PatternFill("solid",fgColor="1F3A4D")
gf={"🟢":PatternFill("solid",fgColor="E2F1EC"),"🔴":PatternFill("solid",fgColor="F3E2E0"),"🟡":PatternFill("solid",fgColor="FBF0DF")}
H=["KB Feature","한글","구분","정면","측면","구현 VF","방법 / 왜 되나·안 되나"]
W=[34,20,16,6,6,10,58]
for j,h in enumerate(H,1):
    c=ws.cell(1,j,h); c.font=Font(name=FONT,bold=True,color="FFFFFF",size=10); c.fill=head
    c.alignment=Alignment(vertical="center",horizontal="center"); c.border=bd
    ws.column_dimensions[get_column_letter(j)].width=W[j-1]
for i,row in enumerate(R,2):
    for j,v in enumerate(row,1):
        c=ws.cell(i,j,v); c.border=bd; c.font=Font(name=FONT,size=10,bold=(j==1))
        c.alignment=Alignment(vertical="center",horizontal="center" if j in(3,4,5,6) else "left",wrap_text=(j==7))
        if j==3: c.fill=gf.get(str(v)[:2][:1]+str(v)[1:2],gf.get(v[0],PatternFill()))
n=sum(1 for r in R if r[2].startswith("🟢"))
ws.freeze_panes="A2"; ws.auto_filter.ref=f"A1:G{len(R)+1}"; ws.sheet_view.showGridLines=False
r=len(R)+3
ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=7)
ws.cell(r,1,f"합계 18: 몸 지표(구현) {n} · 클럽 필요 4 · 손목 보류 4 · 압력판 필요 1. "
            "P2 시점은 프록시(손높이가 골반높이를 넘는 프레임, 불성립 시 보간) — 모든 P2 측정에 interpolated_event 플래그. "
            "몸 지표는 천천히 변해 시점 오차에 강건 / 클럽 방향은 정확한 프레임 필요라 근사 불가.").font=Font(name=FONT,size=9,italic=True,color="666666")
out=os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "docs", "DOH_P2_구현표.xlsx")
wb.save(out); print(f"P2 {len(R)}행 (구현 {n}) saved:", out)
