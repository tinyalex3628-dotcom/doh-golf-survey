# DOH 뷰별 측정가능표 — 시트1: 진단 17규칙 × 정면/측면 / 시트2: 측정 42 × 정면/측면
# 근거 = 엔진 실제 게이팅(wham_golf_metrics.py FRONT/SIDE/BOTH + analyzer2 규칙) — 문서화.
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
FONT="맑은 고딕"

OK="✅ 가능"; LO="◐ 저신뢰"; NO="✕ 안 나옴"; TODO="· 미구현"
# [진단명, 정면, 측면, 왜 (한 줄)]
D=[
("스웨이",OK,NO,"좌우(타깃선) 이동 — 측면에선 화면 깊이축이라 원리적 불가"),
("리버스 피벗",OK,NO,"좌우 체중편향 — 상동"),
("리버스 스파인",OK,NO,"척추 좌우 기울기 — 정면에서만 보임"),
("힙 슬라이드 과다",OK,NO,"골반 좌우 이동"),
("행잉백",OK,NO,"임팩트 좌우 체중 잔류"),
("자세 무너짐 (백스윙)",NO,OK,"척추 앞숙임 변화(시상면) — 정면에선 깊이축"),
("일어남 (임팩트)",NO,OK,"상동"),
("치킨윙 경향",OK,OK,"3D 관절 굽힘각 — 뷰 무관 강건"),
("플라잉 엘보 경향",OK,OK,"팔꿈치 높이(수직) — 양쪽 보임"),
("오버스윙 경향",OK,OK,"리드팔 수직각 — 양쪽"),
("상체 회전 부족",OK,OK,"3D 복원 회전 — 뷰 무관 (절대각 스케일 주의)"),
("X-Factor 이상",OK,OK,"상동"),
("트레일 무릎 펴짐",OK,OK,"3D 무릎 굽힘각"),
("머리 상하 출렁",OK,OK,"수직 이동 — 양쪽"),
("템포 이상",OK,OK,"시간 산술 — 뷰 무관"),
("얼리 익스텐션",LO,OK,"측면=tush line 이탈(볼쪽 축이 화면 안이라 측정 가능) / 정면=벨트버클 상승분만 부분 포착"),
("오버더톱 경향",NO,OK,"손이 볼쪽으로 던져짐 — 측면에선 화면 안 축(깊이 아님)"),
]

# [부위, M, 측정, 정면, 측면, 구현, 비고]
M=[
("골반","M101","골반 회전",OK,OK,"됨","3D 복원 회전"),
("골반","M103","골반 좌우기울기",OK,NO,"됨(표시만)","전두면"),
("골반","M104","골반 좌우이동",OK,NO,"됨","스웨이·슬라이드·행잉백 재료"),
("골반","M106","골반 높이(상승)",OK,OK,"됨","얼리익스텐션 수직성분 = 벨트버클 뜸(VF121)"),
("골반","M102","골반 앞뒤기울기",NO,OK,"안 됨","시상면 (Sportsbox Bend)"),
("골반","M105","골반 앞뒤이동",NO,OK,"됨","tush line 이탈 — 측면에선 화면 안 축"),
("몸통","M201","흉곽(어깨) 회전",OK,OK,"됨",""),
("몸통","M202","척추 앞숙임",NO,OK,"됨","자세각 — 시상면"),
("몸통","M203","척추 좌우기울기",OK,NO,"됨","리버스 스파인 재료"),
("몸통","M204","상체중심 좌우이동",OK,NO,"안 됨",""),
("몸통","M205","상체중심 높이",OK,OK,"안 됨",""),
("몸통","M206","상체 깊이",NO,LO,"안 됨","깊이추정"),
("머리","M301","머리 좌우이동",OK,NO,"됨","스웨이 재료"),
("머리","M302","머리 상하이동",OK,OK,"됨",""),
("머리","M303","머리 깊이",NO,LO,"안 됨","헤드비하인드볼 — 깊이추정"),
("어깨","M401","양 어깨 높이차",OK,NO,"안 됨","전두면"),
("어깨","M402","어깨선 기울기",OK,LO,"됨(표시만)","어깨플레인은 측면 우세·좌우틸트는 정면"),
("어깨","M403","리드 어깨 들림",OK,OK,"안 됨","수직 — 양쪽"),
("어깨","M404","트레일 어깨 들림",OK,OK,"안 됨",""),
("팔","M501","리드 팔꿈치 굽힘",OK,OK,"됨","치킨윙 재료"),
("팔","M502","트레일 팔꿈치 굽힘",OK,OK,"됨(표시만)",""),
("팔","M504","리드팔 플레인 각",NO,OK,"안 됨","플레인 — 측면 전용"),
("팔","M506","트레일 팔꿈치 높이",OK,OK,"됨","플라잉 엘보 재료"),
("팔","M507","리드팔–몸통 간격",OK,OK,"안 됨",""),
("팔","M508","트레일팔–몸통 간격",OK,OK,"안 됨",""),
("팔","M505","트레일 팔꿈치 깊이",NO,LO,"안 됨","깊이추정"),
("손","M604","손 높이",OK,OK,"됨","오버스윙 재료"),
("손","M606","손 좌우(볼 대비)",OK,NO,"안 됨","핸드퍼스트"),
("손","M605","손 깊이",NO,OK,"됨","오버더톱 — 측면에선 화면 안 축"),
("손","M607","손–몸통 거리",NO,LO,"안 됨","깊이 성분 큼"),
("하체","M701","리드 고관절 굽힘",LO,OK,"안 됨","힙 힌지 — 시상면 우세"),
("하체","M702","트레일 고관절 굽힘",LO,OK,"안 됨",""),
("하체","M705","리드 무릎 굽힘",OK,OK,"됨(표시만)",""),
("하체","M706","트레일 무릎 굽힘",OK,OK,"됨","무릎 펴짐 재료"),
("하체","M709","리드 무릎 좌우이동",OK,NO,"안 됨",""),
("하체","M710","트레일 무릎 좌우이동",OK,NO,"안 됨",""),
("하체","M707","리드 무릎 안팎",OK,NO,"안 됨","valgus — 저신뢰 예정"),
("하체","M708","트레일 무릎 안팎",OK,NO,"안 됨",""),
("무게중심","M801","무게중심 좌우",OK,NO,"안 됨","체중이동 대리"),
("무게중심","M802","무게중심 상하",OK,OK,"안 됨",""),
("무게중심","M803","무게중심 깊이",NO,LO,"안 됨","얼리익스텐션 대리"),
("시간","—","템포 (백:다운)",OK,OK,"됨",""),
]

wb=Workbook()
bd=Border(*[Side(style="thin",color="DDE3E6")]*4)
head=PatternFill("solid",fgColor="1F3A4D")
cf={OK:("1F7A67","E2F1EC"),LO:("B5720F","F6ECD8"),NO:("A8443C","F3E2E0")}
def cell_view(ws,r,c,v):
    x=ws.cell(r,c,v); x.border=bd
    fg,bg=cf.get(v,("5D6B78",None))
    x.font=Font(name=FONT,size=10,bold=True,color=fg)
    if bg: x.fill=PatternFill("solid",fgColor=bg)
    x.alignment=Alignment(horizontal="center",vertical="center")

ws=wb.active; ws.title="진단 17 × 뷰"
H=["진단명","정면(FO)","측면(DTL)","왜 그런가"]
W=[22,12,12,52]
for j,h in enumerate(H,1):
    c=ws.cell(1,j,h); c.font=Font(name=FONT,bold=True,color="FFFFFF",size=10); c.fill=head
    c.alignment=Alignment(horizontal="center",vertical="center"); c.border=bd
    ws.column_dimensions[get_column_letter(j)].width=W[j-1]
for i,row in enumerate(D,2):
    c=ws.cell(i,1,row[0]); c.font=Font(name=FONT,size=10,bold=True); c.border=bd
    cell_view(ws,i,2,row[1]); cell_view(ws,i,3,row[2])
    c=ws.cell(i,4,row[3]); c.font=Font(name=FONT,size=10); c.border=bd
    c.alignment=Alignment(vertical="center",wrap_text=True)
ws.freeze_panes="A2"; ws.auto_filter.ref=f"A1:D{len(D)+1}"; ws.sheet_view.showGridLines=False
nFO=sum(1 for d in D if d[1]==OK); nDT=sum(1 for d in D if d[2]==OK); nDL=sum(1 for d in D if d[2]==LO)
r=len(D)+3
for txt in [
 f"요약: 정면(FO) = 진단 {nFO}/17 가능 · 측면(DTL) = {nDT}/17 + 저신뢰 {nDL} = 최대 {nDT+nDL}/17.",
 "핵심: 좌우 계열(스웨이·리버스·슬라이드·행잉백) 5개=정면 전용 / 앞뒤 계열(자세각·일어남·얼리익스텐션·OTT) 4개=측면 우세. "
 "※ 검수3차 정정: 측면에서 '앞뒤'는 화면 안 축(깊이축은 타깃선) → tush line·손깊이는 저신뢰가 아니라 정상 측정. 얼리익스텐션은 정면에서도 벨트버클 상승분으로 부분 포착.",
 "이 표는 엔진에 이미 내장된 자동 처리의 문서화 — 안 나오는 항목은 값이 null이 되고 화면에 '판정불가/측정 불가'로 뜸(지어내지 않음).",
 "권장: 제대로 보려면 정면+측면 각 1회(같은 스윙 아니어도 됨). 한 번만 찍는다면 — 회전·팔·무릎은 어느 쪽이든 나오므로, 보고 싶은 결함 쪽 각도로.",
]:
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=4)
    ws.cell(r,1,txt).font=Font(name=FONT,size=9,italic=True,color="666666"); r+=1

ws2=wb.create_sheet("측정 42 × 뷰")
H2=["부위","M코드","측정","정면(FO)","측면(DTL)","구현","비고"]
W2=[9,8,20,12,12,11,30]
for j,h in enumerate(H2,1):
    c=ws2.cell(1,j,h); c.font=Font(name=FONT,bold=True,color="FFFFFF",size=10); c.fill=head
    c.alignment=Alignment(horizontal="center",vertical="center"); c.border=bd
    ws2.column_dimensions[get_column_letter(j)].width=W2[j-1]
for i,row in enumerate(M,2):
    for j,v in enumerate([row[0],row[1],row[2]],1):
        c=ws2.cell(i,j,v); c.font=Font(name=FONT,size=10,bold=(j==2)); c.border=bd
        c.alignment=Alignment(vertical="center",horizontal="center" if j in(1,2) else "left")
    cell_view(ws2,i,4,row[3]); cell_view(ws2,i,5,row[4])
    c=ws2.cell(i,6,row[5]); c.font=Font(name=FONT,size=10,color="1F7A67" if row[5].startswith("됨") else "8B97A5"); c.border=bd
    c.alignment=Alignment(horizontal="center",vertical="center")
    c=ws2.cell(i,7,row[6]); c.font=Font(name=FONT,size=9,color="666666"); c.border=bd
    c.alignment=Alignment(vertical="center",wrap_text=True)
ws2.freeze_panes="A2"; ws2.auto_filter.ref=f"A1:G{len(M)+1}"; ws2.sheet_view.showGridLines=False
fo_ok=sum(1 for m in M if m[3]==OK); dt_ok=sum(1 for m in M if m[4]==OK); dt_lo=sum(1 for m in M if m[4]==LO)
r=len(M)+3
ws2.merge_cells(start_row=r,start_column=1,end_row=r,end_column=7)
ws2.cell(r,1,f"요약: 정면 ✅{fo_ok}/42 · 측면 ✅{dt_ok}+◐{dt_lo}/42. ◐=깊이추정(측면에서만, 오차 큼·2군). "
             "'구현 안 됨' 항목도 뷰 판정은 확정 — 엔진에 추가되는 즉시 이 표 그대로 적용.").font=Font(name=FONT,size=9,italic=True,color="666666")
out="/tmp/claude-0/-home-user-doh-golf-survey/22e3c218-6758-56bc-9d04-15083f8a3bd1/scratchpad/DOH_뷰별_측정가능표.xlsx"
wb.save(out)
print(f"진단 {len(D)} (FO {nFO} / DTL {nDT}+{nDL}) · 측정 {len(M)} (FO {fo_ok} / DTL {dt_ok}+{dt_lo})")
print("saved:",out)
