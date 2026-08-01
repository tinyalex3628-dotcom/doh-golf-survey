#!/usr/bin/env python3
"""
KB 인과그래프 추출기 — kb/source/DOH_P*.xlsx → kb/kb_causal_graph.json
===================================================================
사용자가 수작업으로 만든 코칭 지식베이스(엑셀)에서 **인과 엣지**를 뽑아
엔진이 읽을 JSON으로 만든다.

왜 스크립트로 만드나 (2026-07-28)
---------------------------------
기존 `kb_causal_graph.json`은 어느 세션에서 즉석 파싱으로 만들어졌고 **스크립트가
남지 않았다.** 그 결과 누락을 검증할 방법이 없었고, 실제로 두 군데가 비어 있었다:
  · **P2 — 18개 중 4개 누락** (Pelvic Sway / Spine Angle Loss / Lead Knee Flexion / Lead Knee Collapse)
  · **P6 — 21개 전부 누락** (파일 하나가 통째로. 시트 구조가 달라 파서가 못 읽음)
둘 다 '내용이 없어서'가 아니라 **파싱 실패**였다. 그래서 재현 가능한 스크립트로 고정한다.
`tools/README.md`의 원칙과 동일: 수치를 고칠 땐 산출물이 아니라 이 스크립트를 고친다.

두 가지 시트 레이아웃
--------------------
① P1/P2/P4/P7 — 한 Feature의 인과체인 전부가 **한 셀 안**에 줄바꿈으로 들어있다.
② P6         — 체인 **하나당 한 행**. 섹션 머리글(A열)은 첫 행에만 있고 이어지는
               행은 A열이 비어 있다. ①만 가정한 파서는 P6를 통째로 놓친다.

체인 표기법 (양쪽 공통)
----------------------
    Limited Trail Arm Flexibility (Mobility Ref)
    ↓
    Reduced Trail Arm Flexibility During Takeaway.
    (테이크어웨이에서 트레일 팔이 자연스럽게 접히기 어렵다.)
    ↓
    Outside Takeaway (P2)
→ 원소가 [노드, 메커니즘, 노드, 메커니즘, …] 로 번갈아 온다고 보고
  (e[i], e[i+2]) 를 엣지로, 사이의 e[i+1] 을 `mech` 으로 방출한다.
  ↓ 주변의 빈 줄은 표기 흔들림이라 무시한다(이게 P2 누락 4건의 원인이었다).

실행: `python3 tools/build_kb_graph.py`   (openpyxl 필요)
"""
import os, re, json, sys
from openpyxl import load_workbook

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(ROOT, "kb", "source")
OUT = os.path.join(ROOT, "kb", "kb_causal_graph.json")

PHASES = ["P1", "P2", "P4", "P6", "P7"]
ARROW = "↓"

# 섹션 머리글(A열) → 엣지 방향. 원인=이 Feature를 만드는 것 / 예측=이 Feature가 만드는 것
CAUSE_HDR = ("원인 추정", "원인추정", "Root Cause")
PREDICT_HDR = ("이후 나타날 가능성", "Prediction", "예측")
# 인과가 아닌 섹션 — 읽지 않는다(동반발생·검증포인트·사고확장 등은 엣지가 아님)
SKIP_HDR = ("정의", "Strong Co-oc", "추가 검증", "사고확장", "Definition", "비고", "Note")

CONF_RE = re.compile(r"\[\s*(High|Medium|Low|Pattern\s*Dependent)\s*(?:Confidence)?\s*\]", re.I)
BRANCH_RE = re.compile(r"\bBranch\s*([ABC])\b", re.I)


def norm_conf(s):
    s = re.sub(r"\s+", " ", s.strip()).title()
    return "Pattern Dependent" if s.lower().startswith("pattern") else s


def clean(s):
    return re.sub(r"[ \t]+", " ", str(s).replace(" ", " ")).strip()


def parse_chunk_elements(lines):
    """빈 줄로 나뉜 한 덩어리(lines) → ↓ 기준 원소 리스트.
    한 원소가 여러 줄일 수 있다(영문 문장 + 다음 줄 한글 괄호) → 공백으로 합친다."""
    elems, cur = [], []
    for ln in lines:
        if ln.strip() == ARROW:
            elems.append(" ".join(cur).strip()); cur = []
        else:
            cur.append(ln.strip())
    elems.append(" ".join(cur).strip())
    return [e for e in elems if e]


def is_mech(e):
    """원소가 노드명인가 메커니즘 문장인가.
    KB 표기 관찰: 노드명은 마침표 없이 끝난다('Outside Takeaway (P2)',
    'Hanging Back (체중 뒤잔류)'), 메커니즘은 문장이라 '.'로 끝나거나
    '(…다.)' 한글 번역 괄호로 끝난다. ↓ 표기가 흔들려 노드-메커니즘 교대가
    깨진 체인에서, 교대 가정 대신 이 분류로 엣지를 세운다."""
    s = e.strip()
    if s.endswith("."):
        return True
    if s.endswith(")"):
        inner = s[s.rfind("(") + 1:-1].strip()
        if inner.endswith("다.") or inner.endswith("."):
            return True
    return False


def chain_edges(elems):
    """원소열 → (src, tgt, mech) 목록. 연속 노드 사이의 메커니즘(들)을 합쳐 mech로.
    메커니즘 없이 노드가 연달아 오면 mech=None 엣지."""
    out, prev_node, mechs = [], None, []
    for e in elems:
        if e.lstrip().startswith(("•", "·")):          # 불릿 = 병렬 결과
            b = e.lstrip("•·-— ").strip()
            if prev_node and b:
                out.append((prev_node, b, " ".join(mechs) or None))
            continue
        if is_mech(e):
            mechs.append(e); continue
        if prev_node is not None:
            out.append((prev_node, e, " ".join(mechs) or None))
        prev_node, mechs = e, []
    return out


def parse_section(text, owner, phase, kind):
    """섹션 본문 → 엣지 리스트."""
    if not text:
        return []
    # ↓ 앞뒤의 빈 줄 제거 — 이 표기 흔들림이 P2 4건을 통째로 날렸던 원인
    raw = [clean(l) for l in str(text).split("\n")]
    lines, i = [], 0
    while i < len(raw):
        ln = raw[i]
        if ln == "":
            # 다음 유효 줄이 ↓ 거나, 직전이 ↓ 면 이 빈 줄은 구분자가 아니다
            nxt = next((raw[j] for j in range(i + 1, len(raw)) if raw[j] != ""), None)
            prev = lines[-1] if lines else None
            if nxt == ARROW or prev == ARROW:
                i += 1; continue
        lines.append(ln); i += 1

    edges, conf, branch = [], "High", None
    chunk = []

    def flush():
        if not chunk:
            return
        for s, t, m in chain_edges(parse_chunk_elements(chunk)):
            edges.append(dict(src=s, tgt=t, mech=m))
        chunk.clear()

    for ln in lines:
        m = CONF_RE.search(ln)
        if m and len(ln) < 60:          # 신뢰도 머리글 줄
            flush(); conf = norm_conf(m.group(1)); continue
        mb = BRANCH_RE.search(ln)
        if mb and len(ln) < 60:
            flush(); branch = mb.group(1).upper(); continue
        if ln == "":
            flush(); continue
        chunk.append(ln)
    flush()

    out = []
    for e in edges:
        if not e["src"] or not e["tgt"] or e["src"] == e["tgt"]:
            continue
        if len(e["src"]) > 200 or len(e["tgt"]) > 200:
            continue
        out.append(dict(ph=phase, owner=owner, kind=kind, conf=conf,
                        src=e["src"], tgt=e["tgt"], mech=e["mech"], branch=branch))
    return out


def sheet_for(wb, phase):
    for cand in (phase, phase.lower(), phase.upper()):
        if cand in wb.sheetnames:
            return wb[cand]
    return wb[wb.sheetnames[0]]


def extract(phase):
    wb = load_workbook(os.path.join(SRC, f"DOH_{phase}.xlsx"))
    ws = sheet_for(wb, phase)
    edges, owner, section, buf = [], None, None, []

    def close_section():
        if owner and section and buf:
            edges.extend(parse_section("\n".join(buf), owner, phase, section))
        buf.clear()

    for row in ws.iter_rows(values_only=True):
        a = clean(row[0]) if row[0] is not None else ""
        b = row[1] if len(row) > 1 and row[1] is not None else None
        if a == "●":                                   # 새 Feature 시작
            close_section(); section = None
            owner = clean(str(b).split("·")[0]) if b else None
            continue
        if a.startswith("■"):                          # 새 섹션
            close_section()
            hdr = a.lstrip("■ ").strip()
            if any(h in hdr for h in CAUSE_HDR):     section = "cause"
            elif any(h in hdr for h in PREDICT_HDR): section = "predict"
            elif any(h in hdr for h in SKIP_HDR):    section = None
            else:                                    section = None
            if section and b: buf.append(str(b))
            continue
        # A열이 비어있고 B열만 있는 행 = 직전 섹션의 이어지는 내용 (P6 레이아웃)
        if not a and b is not None and section:
            buf.append("")          # 행 경계 = 체인 구분자
            buf.append(str(b))
    close_section()
    return edges


V1 = os.path.join(ROOT, "kb", "kb_causal_graph.v1.json")


def main():
    all_edges, report = [], []
    for p in PHASES:
        fp = os.path.join(SRC, f"DOH_{p}.xlsx")
        if not os.path.exists(fp):
            report.append((p, None, "원본 파일 없음")); continue
        e = extract(p)
        all_edges.extend(e)
        report.append((p, len({x['owner'] for x in e}), f"엣지 {len(e)}"))

    # ── 병합 정책 (보수적 패치) ─────────────────────────────────────────
    # v1(kb_causal_graph.v1.json) = 2026-07-25 세션의 원 추출본. 노드명에
    # [예측·검증필요] 주석·P2~P4 교차구간 등 이 파서가 안 내는 표기가 있어
    # 전면 재추출로 갈아끼우면 그 62개 엣지가 유실된다.
    # → v1은 전부 보존하고, v1에 **owner(Feature)가 아예 없던 것**만 추가한다.
    #   (= P2 누락 4개 + P6 통누락 21개. 파싱 실패로 빠졌던 자리만 메움.)
    if os.path.exists(V1):
        v1 = json.load(open(V1, encoding="utf-8"))
        v1_owner_names = set()
        for e in v1["edges"]:
            v1_owner_names.add(re.sub(r"\s*\(P\d+\)\s*$", "", e["owner"]).strip())

        def known(owner):
            base = re.sub(r"\s*\(P\d+\)\s*$", "", owner)
            base = base.split("(")[0].strip()          # P6식 'Name (한글)' → Name
            return any(base and base in o for o in v1_owner_names)

        # 병합 정책 (2026-07-29 수정 — 이전엔 너무 보수적이었다)
        # 예전: v1에 owner가 없는 Feature의 엣지만 추가 → 원본의 51%가 누락됐다.
        #       (v1에 이미 있는 Feature라도 v1이 그 Feature의 엣지를 다 담진 않았음)
        # 지금: **원본에 있는데 v1에 없는 (src,tgt) 쌍은 전부 추가**한다.
        #       v1 엣지는 여전히 하나도 안 버린다(고유 표기 보존).
        v1_pairs = {(e["src"].strip(), e["tgt"].strip()) for e in v1["edges"]}

        def clean_pair(e):
            a, b = e["src"].strip(), e["tgt"].strip()
            if not (3 <= len(a) <= 110 and 3 <= len(b) <= 110):
                return False          # 파싱 잔여물 방어
            return True

        patch = [e for e in all_edges
                 if clean_pair(e) and (e["src"].strip(), e["tgt"].strip()) not in v1_pairs]
        seen, uniq_patch = set(), []
        for e in patch:
            k = (e["ph"], e["kind"], e["src"], e["tgt"])
            if k in seen: continue
            seen.add(k); uniq_patch.append(e)
        edges = v1["edges"] + uniq_patch
        mode = f"병합: v1 {len(v1['edges'])} 전부 보존 + 원본 누락분 {len(uniq_patch)} 추가"
    else:
        seen, edges = set(), []
        for e in all_edges:
            k = (e["ph"], e["kind"], e["src"], e["tgt"])
            if k in seen: continue
            seen.add(k); edges.append(e)
        mode = "전체 추출 모드 (v1 없음)"

    nodes = sorted({e["src"] for e in edges} | {e["tgt"] for e in edges})
    data = {"nodes": nodes, "edges": edges}
    json.dump(data, open(OUT, "w", encoding="utf-8"), ensure_ascii=False, indent=None)

    print("KB 인과그래프 추출 —", mode)
    for p, owners, msg in report:
        print(f"  {p}: Feature(owner) {owners} · {msg}")
    import collections
    print(f"  → 노드 {len(nodes)} · 엣지 {len(edges)}")
    print("     conf:", dict(collections.Counter(e['conf'] for e in edges)))
    print("     kind:", dict(collections.Counter(e['kind'] for e in edges)))
    print("  saved:", OUT)


if __name__ == "__main__":
    main()
