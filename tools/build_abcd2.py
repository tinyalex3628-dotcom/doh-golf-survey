# DOH — KB 인과추적 작업표 v2 (조사기반) : 0.벤치마크 / A / B / C / D
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
FONT="맑은 고딕"; bd=Border(*[Side(style="thin",color="DDE3E6")]*4)
HEAD=PatternFill("solid",fgColor="1F3A4D")
GR=PatternFill("solid",fgColor="E2F1EC"); YE=PatternFill("solid",fgColor="F6ECD8")
RE=PatternFill("solid",fgColor="F3E2E0"); BL=PatternFill("solid",fgColor="E6EDF4")

def sheet(wb,title,H,W,rows,fills=None,first=False,notes=None):
    ws=wb.active if first else wb.create_sheet(); ws.title=title
    for j,h in enumerate(H,1):
        c=ws.cell(1,j,h); c.font=Font(name=FONT,bold=True,color="FFFFFF",size=10); c.fill=HEAD
        c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c.border=bd
        ws.column_dimensions[get_column_letter(j)].width=W[j-1]
    ws.row_dimensions[1].height=34
    for i,row in enumerate(rows,2):
        for j,v in enumerate(row,1):
            c=ws.cell(i,j,v); c.border=bd; c.font=Font(name=FONT,size=9.5,bold=(j==1))
            c.alignment=Alignment(vertical="center",wrap_text=(W[j-1]>=24),
                                  horizontal="center" if W[j-1]<=11 else "left")
            if fills:
                f=fills(j,v,row)
                if f: c.fill=f
    ws.freeze_panes="A2"; ws.sheet_view.showGridLines=False
    ws.auto_filter.ref=f"A1:{get_column_letter(len(H))}{len(rows)+1}"
    if notes:
        r=len(rows)+3
        for t in notes:
            ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=len(H))
            ws.cell(r,1,t).font=Font(name=FONT,size=9,italic=True,color="666666"); r+=1
    return ws

wb=Workbook()

# ══════ 0. 벤치마크 — 조사 근거 ══════
BM=[
("우리 (DOH)","단일 스마트폰 · NLF(범용 3D 포즈)","포즈 24관절 · 클럽/압력 없음","공개 검증 없음 (자체 합성검증만)",
 "회전 상대값은 재현성 확인(4스윙 −95~−100°). 절대 스케일 미보정",
 "무료 · 자체","—"),
("Sportsbox 3D Golf","단일 스마트폰 · 골프전용 학습(30+ keypoint, 클럽·볼 포함)",
 "6DOF × 골반/흉곽 (Sway·Lift·Thrust·Turn·Bend·Side Bend)",
 "AMM3D(전자기 골드스탠다드) 대조 30스윙",
 "★ 각도 오차 평균 약 2° (어드레스·탑·임팩트) — 우리가 목표로 삼을 상한선",
 "유료 구독","sportsbox.ai/accuracy"),
("Onform","단일 카메라 3D 포즈 (모델카드 공개, 2026-04)","골프 스윙 3D 포즈","모델카드 공개(PDF 접근 차단)",
 "공개 모델카드를 내놓았다는 것 자체가 업계 검증 압력 신호","유료","onform.com 모델카드"),
("HackMotion","손목 IMU 센서 (하드웨어 착용)","손목 굴신/편향/회전",
 "센서 직접 측정","'손목 움직임은 카메라로 분석하기 매우 어렵다'가 업계 공통 인식 — 그래서 센서로 감",
 "유료 하드웨어","hackmotion / 업계 리뷰"),
("GEARS · AMM3D · Vicon","마커 부착 광학/전자기","전신 정밀 6DOF","골드 스탠다드",
 "마커기반 오차 0.5~1.5mm — 모든 비교의 기준","고가 설비","마커/마커리스 비교 리뷰(MDPI Sensors 2025)"),
("2D 영상앱 (일반)","단일 카메라 2D 오버레이","각도선·박스 그리기","—",
 "★ '2D 영상은 원리적으로 회전을 못 잰다' — 선·원·화살표를 아무리 그려도 몸통 회전은 못 읽음",
 "무료~저가","Driveline Baseball 'Pitfalls of 2D'"),
]
sheet(wb,"0. 벤치마크(조사근거)",
 ["엔진","방식","무엇을 재나","검증 방법","정확도 / 핵심 사실","비용","출처"],
 [16,34,30,24,52,12,26],BM,
 lambda j,v,row: (GR if row[0]=="우리 (DOH)" else None) if j==1 else None, first=True,
 notes=[
 "■ 학계 경고 (반드시 인지) — Ingwersen et al., 'Evaluating current state of monocular 3D pose models for golf' (NLDL 2023, DTU):",
 "   범용 단일카메라 3D 포즈 모델을 모션캡처와 정량 비교한 결과 \"운동학 분석에 쓰기엔 오차가 너무 크다(too erroneous for kinematic analysis)\",",
 "   \"현재 모델을 그대로 고급 골프 분석에 쓸 수 없다(cannot be used out of the box)\"고 결론. ← 우리가 지금 하는 방식에 대한 직접 경고.",
 "■ 그러나 해법도 같은 문헌에 있음 — GolfPose(ICPR 2024): 범용 MixSTE의 골프 MPJPE 109.4mm → 골프전용 학습 후 35.6mm(클럽 포함 32.3mm).",
 "   즉 '범용 모델을 그대로 쓰는 것'이 문제이고, 골프 특화가 답. Sportsbox가 2°를 내는 이유도 골프전용 학습 + 클럽/볼 keypoint.",
 "■ 마커리스 일반 정확도: 위치 5~20mm · 관절각 평균 2.31°±4.00° (MDPI Sensors 2025 리뷰). 손/손가락은 8~20°로 급격히 악화 → 손목 제외 근거.",
 "■ 촬영각이 오차의 1차 결정요인: 화면-내(in-plane) 관절각은 정확, 화면-밖(out-of-plane)은 최악 → 우리 뷰-게이팅 설계가 문헌과 일치.",
 "▶ 우리 위치: 정확도는 Sportsbox(2°)와 마커기반(0.5~1.5mm) 사이가 아니라, '검증 안 된 범용 모델' 구간에 있음. 이걸 숨기지 말고 UI에서 '추정'으로 표기하는 게 유일하게 정직한 길.",
])

# ══════ A ══════
A=[
("VF014/017/019","@P3 어깨·골반·X-Factor","P3","✅ 구현완료(커밋 6879451)",
 "회전 시계열에서 P3 프레임만 읽으면 끝 — 기준치 불필요",
 "Sportsbox는 모든 프레임의 6DOF를 연속 제공. 우리도 시계열은 있었으나 P4만 소비 중이었음",
 "동등 (시점 읽기는 난이도 0)","Body Rotation 계열"),
("VF047","Shoulder-Hip Sep @P5","P5","✅ 구현완료",
 "X-Factor Stretch 판정 재료 (P4 대비 증감)",
 "Sportsbox·K-Vest 모두 X-Factor Stretch를 핵심 지표로 판매",
 "동등","X-Factor Stretch"),
("VF151","Thorax Turn @P7 (신설)","P7","✅ 구현+Spec등록",
 "임팩트 흉곽 회전. 기존에 ID 자체가 없었음(골반은 VF075)",
 "Sportsbox 'Chest Turn at Impact'와 동일 개념. 부호규약도 동일(+ = 목표쪽 열림)",
 "동등 — 단 절대 스케일 미보정","Thorax Open/Closed (P7) ★5엣지"),
("VF103/104","@P10 피니시 회전","P10","✅ 구현완료","피니시 회전 완성도",
 "상용앱 대부분 피니시는 정성 평가에 그침 — 우리가 수치화하면 오히려 차별점",
 "동등","Finish 계열"),
("—","@P2 어깨·골반 회전","P2","✕ 불가",
 "P2=샤프트 평행 → 클럽검출 필요. 우리 P구간 검출에 P2 없음",
 "Sportsbox는 클럽 keypoint가 있어 P2 검출 가능 (우리와의 실질 격차)",
 "✕ 클럽검출 2차까지 보류","Body Rotation Dominant (P2) 7엣지 — 보류"),
]
sheet(wb,"A. 완료 — 시점 추가",
 ["VF","측정","시점","상태","기준을 어떻게 잡았나","상용 엔진은 어떻게 하나","우리 구현 가능성 평가","살아나는 KB 노드"],
 [14,24,7,20,38,46,26,28],A,
 lambda j,v,row: (GR if str(v).startswith("✅") else (RE if str(v).startswith("✕") else None)) if j==4 else None,
 notes=[
 "A의 성격: 기준치가 필요 없는 유일한 구간이었다 — 값은 이미 있고 시점만 안 읽고 있었음. feature 29→37개.",
 "⚠️ 단, '값이 나온다 ≠ 판정이 된다'. Thorax Open/Closed(P7) 판정을 켜려면 B의 기준(몇 도부터 열림)이 필요.",
 "▶ 상용 대비 실질 격차는 P2/P6/P8(샤프트 평행) — 전부 클럽 keypoint가 있어야 검출됨. 우리 최대 미싱피스.",
])

# ══════ B ══════ 조사기반 기준표
B=[
("Trail Elbow Flexion @P4","트레일 팔꿈치 내각(180=폄)","< 70° 과굴곡","70~85°","85~115° 정상","> 115° 과신전","B",
 "코칭 통설 '탑 90°가 최대' + 엘리트 실측 114°(카메론 스미스·넬리 코다). 주말골퍼는 90°보다 좁게 접히는 경향",
 "Sportsbox: 팔꿈치는 6DOF 주력 지표가 아님(골반·흉곽 중심). GEARS는 팔 세그먼트 각 제공. "
 "즉 상용도 팔꿈치 절대각 기준은 공개 안 함 → 코칭 컨센서스가 사실상 유일 근거",
 "✅ 가능 — 3D 세그먼트 내각은 마커리스가 가장 잘하는 영역(관절각 2.31°±4.00°). 우리 오차 ±3~5° 추정과 정합",
 "golfjournal / 전욱휴레슨 / MDPI Sensors 2025"),
("Hand Height @P4","손 높이 (트레일 어깨 대비)","어깨선 아래=Low","어깨선±","어깨~머리 정상","머리 위=High","B",
 "'탑에서 손은 어깨 높이 또는 살짝 위' + '트레일 어깨 위' 체크포인트. "
 "⚠️ 버치 하몬: 높은 손을 의도적으로 추구 말 것 → High가 항상 결함 아님",
 "Sportsbox: Hand 6DOF(Sway/Lift/Thrust) 제공하되 '높다/낮다' 판정은 안 함 — 수치만 주고 해석은 코치 몫. "
 "우리도 같은 태도(수치+경향)가 안전",
 "✅ 가능 — 수직축 위치는 화면-내 축이라 두 뷰 다 강건",
 "hackmotion.com / golfdigest(Butch Harmon)"),
("Head Lateral @P1→P4","머리 좌우 이동","> 1인치(2.5cm) 과도","—","≤1인치 정상","—","B",
 "'좌우 1인치 초과 시 아이언 일관성 저하'가 널리 쓰이는 기준. 우리는 스탠스폭 정규화 → 스탠스폭 40cm 가정시 ≈0.06",
 "Sportsbox는 Head를 inch 절대값으로 제공(스케일 추정 필요). 우리는 스탠스폭 비율 → 스케일 추정 불필요하나 "
 "상용 수치와 직접 비교 불가. ★ 환산표를 두면 양쪽 다 대응 가능",
 "✅ 가능 — 정면에서 화면-내 축. 단 절대 inch로 내려면 신장/스탠스 입력 필요",
 "theleftrough.com / sportsbox 정의"),
("Head Vertical @P1→P4","머리 상하 이동","> 1인치 과도","—","≤1인치 정상","—","B",
 "상동. 단 다운스윙의 약간의 head dip은 지면반력상 정상이라는 언급 → 백스윙 구간에 한정 적용 권장",
 "Sportsbox 'Head Lift' 대응. 부호규약 동일(위=+)",
 "✅ 가능 — 수직축은 두 뷰 다 강건","theleftrough.com"),
("Shoulder Turn @P4","어깨(흉곽) 회전","< 80° 부족","80~90°","90~105° 정상","> 105° 과회전","B",
 "교과서 '상체 90°/하체 45°'. 기존 seed(75°)보다 상향 권고",
 "Sportsbox 'Chest Turn' — AMM3D 대조 오차 2°로 검증된 영역. K-Vest도 동일 개념",
 "◐ 조건부 — 측정은 되나 ⚠️우리 절대 스케일 미보정(실측 −109°는 과대 의심). "
 "스케일 보정 전엔 이 기준 적용 금지",
 "hankyung magazine / sportsbox"),
("Hip Turn @P4","골반 회전","< 35° 부족","35~45°","45~55° 정상","> 55° 과회전","B",
 "교과서 하체 45°. KB에서 Pelvis Rotation Excessive는 원인 8엣지 = 최상위 뿌리 중 하나",
 "Sportsbox 'Pelvis Turn' — 부호규약 명시(−=목표 반대/closed, +=목표쪽/open). 우리와 부호 통일 필요",
 "◐ 조건부 — 상동(스케일 보정 선행)","hankyung / sportsbox"),
("Thorax Turn @P7","임팩트 흉곽 회전","< 0° 닫힘","0~15°","15~40° 정상","> 40° 과열림","C",
 "임팩트 흉곽은 약간 열린 상태가 정상이라는 통설은 있으나 수치 문헌 확보 실패 → seed",
 "Sportsbox는 impact 시점 chest turn을 제공하나 정상범위는 비공개(코치가 해석). "
 "AMM3D 검증이 '어드레스·탑·임팩트' 3지점이라 임팩트 정확도 자체는 검증된 구간",
 "◐ 조건부 — A작업으로 값은 이제 나옴. 기준은 실측 누적 필요(1순위)",
 "(수치 문헌 없음 — 실측 필요)"),
("X-Factor @P5 (Stretch)","전환 분리각 변화","P4 대비 감소=실패","유지","P4 대비 증가=정상","—","B",
 "★ 절대값이 아니라 P4→P5 변화 방향으로 판정 → 임계값 불필요. KB도 'Stretch가 큰가 작은가'로 질문",
 "Sportsbox·K-Vest 공히 X-Factor Stretch를 파워 핵심으로 다룸. 방향 판정이라 스케일 오차에 강함",
 "✅ 가능 ★ 모범사례 — 스케일 미보정 상태에서도 쓸 수 있는 유일한 회전 지표",
 "KB P4 원문 + 코칭 컨센서스"),
("Trail Elbow Flexion @P1","어드레스 팔꿈치","< 150° 과굴곡","150~165°","165~180° 정상","—","C",
 "어드레스는 거의 편 상태가 기본. 문헌 수치 확보 실패 → seed",
 "상용도 어드레스 팔꿈치 기준은 공개 안 함","✅ 가능(측정 쉬움) / 기준만 seed",
 "(문헌 없음)"),
("Shoulder Turn @P3","백스윙 중간 회전","< 40° 부족","40~55°","55~70°","—","C",
 "P3 시점 회전 문헌 없음 → P4의 약 60%로 seed",
 "Sportsbox는 연속 그래프로 제공하고 특정 시점 기준은 없음 — 궤적 모양을 코치가 봄. "
 "★ 우리도 '시점 기준'보다 '궤적 형태' 판정이 나을 수 있음",
 "✅ 가능 / 기준 seed","(문헌 없음)"),
]
sheet(wb,"B. 기준 조사(상용 병기)",
 ["대상","측정 (단위·부호)","🔴 이상","🟡 경계","🟢 정상","🟠 반대극","근거\n등급","조사 근거 (문헌·코칭)","상용 엔진은 어떻게 하나","우리 구현 가능성 평가","출처"],
 [22,26,15,11,17,16,6,46,52,44,26],B,
 lambda j,v,row: ({"A":GR,"B":YE,"C":RE}.get(str(v)) if j==7 else
                  (GR if (j==10 and str(v).startswith("✅")) else (YE if (j==10 and str(v).startswith("◐")) else None))),
 notes=[
 "■ 조사에서 확인된 불편한 사실: 팔꿈치·손높이 등의 논문급 정량 데이터는 사실상 공개돼 있지 않다 —",
 "   'Elbow kinematics have rarely been fully described in terms of anatomical joint angle kinematics' (MDPI Sports 2022 리뷰).",
 "   그리고 상용 엔진들도 '정상 범위'를 공개하지 않는다(수치만 주고 해석은 코치). 즉 A등급 근거는 업계 전체에 존재하지 않음.",
 "■ 그래서 현실적 최선은 B등급(코칭 컨센서스)이며, 우리의 차별점은 '기준의 정확도'가 아니라 '기준을 공개하고 근거를 표시하는 정직성'.",
 "■ 회전 3종(어깨·골반·임팩트흉곽)은 ◐ — 측정은 되지만 절대 스케일 보정 전이라 기준 적용 보류. X-Factor Stretch만 방향 판정이라 즉시 사용 가능.",
 "▶ 실행 권고: 스케일 보정 → 회전 기준 확정. 그 전까지는 회전 관련 판정에 [스케일 보정전] 태그 유지.",
])

# ══════ C ══════
C=[
("Hand Depth Deep/Shallow (P4)",11,"손 깊이","기준선 미확정 (흉곽 상대 vs 뒷꿈치 수직선)",
 "① heel line: 양 발목의 지면 수직면 기준, 손중점의 볼-반대방향 변위 ÷ 스탠스폭\n"
 "② 흉곽 상대(KB 원문): 흉곽 회전축 대비 손의 후방 변위",
 "Sportsbox 'Hand Thrust' = 전후 직선이동(inch), 원점은 어드레스. "
 "즉 상용은 ①에 가까운 '고정 기준면 대비 변위'를 쓴다. KB의 흉곽상대는 더 정교하나 회전오차 전파",
 "✅ 가능(측면) — 측면에서 앞뒤축은 화면-내. ①은 발이 고정이라 안정적. "
 "권장: ①을 주판정, ②를 보조. 상용과 정의가 맞아 비교도 쉬움",
 "sportsbox 6DOF 정의 / 사용자 현장지식(heel line)"),
("Pelvis Tilt Anterior/Posterior (P1)",9,"골반 앞뒤기울기(M102)","설계만(2군 예약칸)",
 "골반중심→척추1 벡터의 시상면 기울기(up축 기준). SMPL 관절만으론 골반 자체 틸트가 약해 상체 기울기와 분리 필요",
 "★ Sportsbox 'Pelvis Bend'가 정확히 이것 — 6DOF 정식 지표로 판매 중이고 AMM3D 대조 2° 검증 구간에 포함",
 "◐ 조건부 — 측면 필요 + SMPL 골반 표현력 한계. Sportsbox는 골프전용 학습으로 해결, 우리는 범용 NLF라 열세 예상. "
 "구현하되 신뢰도 낮게 표기 권장",
 "sportsbox 6DOF / SMPL 구조 한계"),
("Shoulder Height Trail High (P1)",5,"양 어깨 높이차(M401)","✅ 구현 완료(2026-07-28) — VF008",
 "(trail_sh − lead_sh)·up ÷ 어깨너비. 오른손잡이는 트레일 어깨가 자연히 낮은 게 정상 → 부호 주의",
 "Sportsbox 'Chest Side Bend'(+ = 트레일측이 낮아짐)가 유사 개념 — 다만 각도(°), 우리는 비율. "
 "각도로 내면 상용과 직접 비교 가능",
 "✅ 구현됨 — VF008 @P1, 단위 deg(Sportsbox Side Bend 대조 가능), 정면 전용. "
 "판정 구간 정상≥5°/주의2~5/과다0~2/심함≤0 (근거등급 B). "
 "selfcheck_metrics.py: 기하 정답 대조 오차 0.00°, up축 전두면 롤 오차만 1:1 전파",
 "sportsbox Side Bend 정의"),
("Lead Arm Plane Flat/Steep (P4)",7,"리드팔 플레인각(M504)","설계만",
 "리드 어깨→리드 손목 벡터와 지면 사이 각",
 "GEARS 등은 'Arm Plane'을 클럽 플레인과 함께 제공. Sportsbox 6DOF엔 팔 플레인 별도 항목 없음 "
 "→ 상용도 클럽 없이는 팔 플레인만 다루는 편",
 "✅ 가능(측면 우세) — 클럽 없이 팔만으로도 정의 성립. KB의 Hand Depth와 공통원인이라 함께 구현시 시너지",
 "KB P4 원문(Arm Vertical/Horizontal Motion 공통원인)"),
("Head Height High (P4)",3,"머리 높이","(재분류) 측정 이미 있음",
 "VF033(머리 상하이동)이 이미 있음 — 'High' 판정은 그 값의 +방향 기준만 붙이면 됨",
 "Sportsbox 'Head Lift'","✅ 즉시 — 실은 C가 아니라 B. 기준만 필요","—"),
("Trail/Lead Knee 계열 (P4)",4,"무릎각","(재분류) 측정 이미 있음",
 "VF039/040/088 존재 — 기준 구간만 없음",
 "Sportsbox는 무릎을 6DOF 주력에서 제외(골반·흉곽 중심)","✅ 즉시 — B로 이관","—"),
]
sheet(wb,"C. 측정 구현 설계",
 ["KB 노드","엣지","무엇","현재 왜 안 되나","측정 방법 제안 (계산식)","상용 엔진은 어떻게 하나","우리 구현 가능성 평가","근거/출처"],
 [28,6,16,26,52,52,50,30],C,
 lambda j,v,row: (GR if (j==7 and str(v).startswith("✅")) else (YE if (j==7 and str(v).startswith("◐")) else None)),
 notes=[
 "▶ 우선순위(효과÷난이도): ①~③ 진행 상황 — ③ Hand Depth ✅구현(VF028, 2026-07-28, 판정구간은 측면 실측 대기) ·",
 "   ① 어깨 높이차 ✅구현+판정구간까지(VF008, 2026-07-28) · 남은 것 → ② 리드팔 플레인(7엣지, 클럽 불필요) → ④ 골반 틸트(범용모델 열세 예상)",
 "★ Hand Depth 기준선 권고: heel line(①)을 주판정으로. 이유 — (a)발이 고정이라 회전오차 전파 없음 (b)Sportsbox Thrust와 정의가 사실상 동일해 비교·검증 가능 (c)사용자 현장 관행과 일치.",
 "   KB의 흉곽상대(②)는 '회전 대비 과한가'를 볼 때 보조 지표로 병행.",
 "■ 2건(Head Height·Knee)은 재분류 — 측정이 이미 있어 C가 아니라 B(기준만 필요).",
])

# ══════ D ══════
D=[
("Standing Up (P7)",4,"임팩트에 몸이 일어남","측정 방법이 3가지로 갈림",
 "① 척추각 증가(VF076, 구현됨) ② 골반 높이 상승(M106) ③ 머리 높이 상승(M302)",
 "Sportsbox는 'Pelvis Lift'(수직 상승)와 'Chest Bend'(전방굴곡 감소)를 별도 제공 — 즉 상용은 ②와 ①을 나눠서 봄. "
 "합치지 않고 둘 다 보여주는 방식",
 "✅ 이미 ①로 구현('일어남' 규칙). ②는 M106 추가하면 됨 — 상용처럼 둘 다 내는 것 권장",
 "★ 확인: KB의 Standing Up = 우리 '일어남'과 같은 개념인가?"),
("Early Extension vs Standing Up",4,"둘 다 '일어남' 계열","KB에 둘 다 존재 — 경계 불명",
 "통설: EE = 골반이 볼쪽 전진(수평·tush line 이탈) / SU = 수직 상승 위주",
 "★ 조사결과 명확: tush line = 어드레스에서 엉덩이 뒤로 그은 수직선, 다운스윙에 이 선을 잃으면 EE. "
 "Sportsbox는 EE를 'Pelvis Thrust'(전후)로, SU를 'Pelvis Lift'(상하)로 축 분리",
 "✅ 우리도 이미 분리돼 있음 — EE=M105(볼쪽 이동), SU=VF076(척추각). 상용과 같은 구조. "
 "다만 우리 EE는 '변위'만, tush line 기준면 개념을 넣으면 더 정확",
 "rotaryswing / bradleyhughesgolf / sportsbox"),
("Side Bend Excessive (P7)",3,"임팩트 측면굴곡","우리 척추 좌우기울기와 같은 건가",
 "M203(척추 좌우기울기) P7 시점 = A작업과 동형(값 있음, 시점만)",
 "★ Sportsbox 'Side Bend' 정의 확보: 전후축(Z축) 둘레 회전, + = 트레일측이 리드측보다 낮아짐. "
 "임팩트 트레일 사이드밴드는 정상 스윙의 특징이기도 함 → '과도'만 문제",
 "✅ 가능 — M203 P7 시점 추가하면 즉시. 부호규약을 Sportsbox와 통일 권장(+=트레일 낮음)",
 "sportsbox 6DOF 정의"),
("Forward Bend Excessive (P7)",2,"임팩트 전방굴곡","척추 앞숙임과 같은 건가",
 "M202(척추 앞숙임) — VF076이 이미 P1→P7 변화를 봄",
 "Sportsbox 'Bend' = 전후 기울기 각도. 우리 척추 앞숙임과 동일 개념",
 "✅ 가능 — 변화량(VF076) 말고 P7 절대각도 함께 내면 KB 요구 충족",
 "sportsbox 6DOF"),
("Hand Direction Forward (P7)",2,"임팩트 손 위치","'앞'의 기준 불명",
 "① 볼 기준(핸드 퍼스트) ② 어드레스 손 위치 대비 ③ 리드 어깨 수직선 대비",
 "Sportsbox 'Hand Sway/Thrust'는 어드레스 원점 기준(②). 코칭 통념의 핸드퍼스트는 볼 기준(①). "
 "둘이 다름 — 상용은 ②, 코칭은 ①",
 "◐ 조건부 — ②는 즉시 가능(M606), ①은 볼 위치 검출 필요(클럽/볼 2차)",
 "★ 확인: ②(어드레스 대비)로 시작해도 되나? ①은 볼 검출 후"),
("Inside/Outside Takeaway (P2)",4,"테이크백 클럽 경로","기준선 + P2 검출 둘 다 없음",
 "① 타깃선 ② 발끝선 ③ 어드레스 샤프트 연장선",
 "상용은 전부 클럽 keypoint로 해결(Sportsbox 30+ keypoint에 클럽 포함). "
 "클럽 없이 손 경로로 근사하는 상용 사례는 확인 못 함",
 "✕ 보류 권장 — P2 미검출 + 기준선 미확정 이중 문제. 손 경로 근사는 신뢰도만 깎음",
 "sportsbox keypoint 구성"),
]
sheet(wb,"D. 정의 확인 필요",
 ["KB 노드","엣지","무엇","왜 애매한가","가능한 정의 (옵션)","상용 엔진은 어떻게 하나 (조사결과)","우리 구현 가능성 평가","확인 요청 / 근거"],
 [26,6,18,24,44,54,50,38],D,
 lambda j,v,row: (GR if (j==7 and str(v).startswith("✅")) else (YE if (j==7 and str(v).startswith("◐")) else (RE if (j==7 and str(v).startswith("✕")) else None))),
 notes=[
 "■ 조사로 4건이 해소됨: Standing Up·Early Extension·Side Bend·Forward Bend는 Sportsbox 6DOF 정의와 1:1 대응이 확인됐고,",
 "   우리 측정도 이미 존재. 축 분리 방식(EE=Thrust 전후 / SU=Lift 상하)도 우리 구조와 동일 → 그대로 채택 권장.",
 "■ 남은 실질 미해결은 2건: Hand Direction Forward(기준 ①볼 vs ②어드레스 — 상용은 ②) / Inside-Outside Takeaway(클럽 필요, 보류).",
 "★ 부호규약 통일 권고: Sportsbox 규약(Turn +=목표쪽 열림 / Side Bend +=트레일측 낮음 / Lift +=위 / Thrust +=앞)을 그대로 채택하면",
 "   나중에 상용 데이터와 대조검증할 때 변환이 필요 없다. 지금 정하는 게 비용 0.",
])
out=os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "docs", "DOH_ABCD_작업표.xlsx")
wb.save(out); print("saved:",out)
print(f"시트: 0.벤치마크 {len(BM)} · A {len(A)} · B {len(B)} · C {len(C)} · D {len(D)}")
